import streamlit as st
import pandas as pd
import openpyxl as op

pagina = st.set_page_config(page_title='Terceiro',
                    layout="wide")

caminho_arquivo = r"C:\Users\arthu\OneDrive\Documentos\Repositórios\projeto_usi_terc\Controle de Terceiros 2024 Atualizada copia - Copia.xlsx"
nome_planilha = 'TABELA UNIFICADA 2024'
nome_planilha2 = 'INSERIR DADOS'

@st.cache_data
def carregar_planilha1():
    df = pd.read_excel(caminho_arquivo, sheet_name=nome_planilha, skiprows=1)
    df['COLETA'] = df['COLETA'].dt.strftime('%d/%m/%Y')
    df['ENTREGA REALIZADA'] = df['ENTREGA REALIZADA'].dt.strftime('%d/%m/%Y')
    df['ENTREGA PREVISTA'] = df['ENTREGA PREVISTA'].dt.strftime('%d/%m/%Y')
    return df

@st.cache_data
def carregar_planilha2():
    return pd.read_excel(caminho_arquivo, sheet_name=nome_planilha2)

df = carregar_planilha1()
df2 = carregar_planilha2()

def encontrar_primeira_linha_vazia(worksheet):
    for row in range(1, worksheet.max_row + 1):  
        if worksheet.cell(row=row, column=1).value is None:
            return row
    return worksheet.max_row + 1

st.title("Usinagem de Terceiro")

try:
    df['OP'] = df['OP'].astype(int) 
except:
    df['OP'] = df['OP'].astype(str) 

try:
    df['DIÂMETRO FP'] = df['DIÂMETRO FP'].astype(int) 
except:
    df['DIÂMETRO FP'] = df['DIÂMETRO FP'].astype(str) 



tab1, tab2, tab3, tab4 = st.tabs(["Entregas em Aberto", "Adicionar Entregas", "Requisitar O.C", "Entregas Realizadas"])

with tab1:
    st.subheader("Entregas em Aberto")
    fornecedor_col = 'FORNECEDOR'

    
    if fornecedor_col in df.columns:
        fornecedores_ordenados = sorted(df[fornecedor_col].dropna().astype(str).unique())
        fornecedores_ordenados.insert(0, "TODOS")
        escolha = st.selectbox("Fornecedor:", fornecedores_ordenados)
        if escolha == "TODOS":
            df_fornecedor = df.dropna(subset=[fornecedor_col], how="all")
        else:
            df_fornecedor = df[df[fornecedor_col].astype(str) == escolha]
        entrega_realizada_col = 'ENTREGA REALIZADA'
        if entrega_realizada_col in df_fornecedor.columns:
            entregas_em_aberto = df_fornecedor[df_fornecedor[entrega_realizada_col].isna()]
            st.write(entregas_em_aberto)
        else:
            st.error(f"A planilha não possui a coluna '{entrega_realizada_col}'.")
    else:
        st.error(f"A planilha não possui a coluna '{fornecedor_col}'.")

with tab4:
    st.subheader("Entregas Realizadas")
    entregas_realizadas_col = 'ENTREGA REALIZADA'
    pesquisa = st.text_input('Pesquisar Item:')
    resultado_pesquisa = df[df['ITEM'].str.contains(pesquisa, case=False, na=False)]
    if pesquisa:
        if not resultado_pesquisa.empty:
            st.write(resultado_pesquisa)
        else:
            st.error('Nenhum resultado encontrado.')
    else:
        if entrega_realizada_col in df.columns:
            df_entrega_realziada = df[df[entregas_realizadas_col].notnull()]
            st.write(df_entrega_realziada)

with tab2:
    st.subheader("Adicionar Entrega")
    
    col1, col2,col3 = st.columns([1,1,1])

    with col1:
        fornecedores_ordenados = sorted(df['FORNECEDOR'].dropna().astype(str).unique())
        escolha_fornecedor = st.selectbox("Fornecedor:", fornecedores_ordenados)
        op_numero = st.text_input("Digite o número da OP:")
        item_codigo = st.text_input("Digite o código do item:")
        item_nf_codigo = st.text_input("Digite o código NF do item:")

    with col2:
        try:
            diametro_fp = st.number_input("Digite o diâmetro do FP:", min_value=0.00)
        except:
            st.error("Por favor não digite valores negativos")
            diametro_fp = 0.00
        qtd_pecas = st.number_input("Digite a quantidade:", min_value=0, max_value=10000)
        obs = st.text_input("Digite a observação:")
        preco = st.number_input("Digite o preço:", min_value=0.00, format="%.2f")
        if preco < 0:
            st.error("O preço não pode ser negativo.")
        buchas_ordenadas = sorted(df['TIPO DE BUCHA'].dropna().astype(str).unique())
        escolha_bucha = st.selectbox("Selecione o tipo de bucha:", buchas_ordenadas)

    with col3:
        operacao_col = 'OPERAÇÃO'
        if operacao_col in df.columns:
            operacao_ordenados = sorted(df[operacao_col].dropna().astype(str).unique())
            operacao_ordenados.insert(0, "ADICIONAR OPERAÇÃO")
            escolha = st.selectbox("Operação:", operacao_ordenados)
            if escolha == "ADICIONAR OPERAÇÃO":
                operacao = st.text_input("Digite a operação:")
            else:
                operacao = escolha
        qtd_fix = int(st.number_input("Selecione a quantidade de FF:", min_value=0, max_value=100))
        data_coleta = st.date_input("Selecione a data da coleta:", format="DD/MM/YYYY")
        data_prevista = st.date_input("Selecione a data de entrega prevista:", format="DD/MM/YYYY")
        
        if st.button('Adicionar Entrega'):
            campos_obrigatorios = [
            ('FORNECEDOR', escolha_fornecedor),
            ('OP', op_numero),
            ('ITEM', item_codigo),
            ('ITEM NF', item_nf_codigo),
            ('QTD', qtd_pecas),
            ('TIPO DE BUCHA', escolha_bucha),
            ('OPERAÇÃO', operacao),
            ('COLETA', data_coleta),
            ('ENTREGA PREVISTA', data_prevista)
                                                ]
            campos_preenchidos = all(campo_valor and str(campo_valor).strip() for campo_nome, campo_valor in campos_obrigatorios) 
            if campos_preenchidos:
                try:
                    nova_linha = {'FORNECEDOR': escolha_fornecedor, 
                                'OP': op_numero,
                                'ITEM': item_codigo, 
                                'ITEM NF': item_nf_codigo,
                                'DIÂMETRO FP': diametro_fp,
                                'QTD': qtd_pecas,
                                'OBS': obs,
                                'VALOR': preco,
                                '': 0,
                                'TIPO DE BUCHA': escolha_bucha,
                                'OPERAÇÃO': operacao,
                                'QTD FF': qtd_fix,
                                'COLETA': data_coleta,
                                'ENTREGA PREVISTA': data_prevista}
                    workbook = op.load_workbook(caminho_arquivo)
                    worksheet = workbook[nome_planilha2]
                    primeira_linha_vazia = encontrar_primeira_linha_vazia(worksheet)
                    for col_num, (key, value) in enumerate(nova_linha.items(), start=1):
                        worksheet.cell(row=primeira_linha_vazia, column=col_num, value=value)
                    workbook.save(caminho_arquivo)
                    st.success('Entrega adicionada')
                except Exception as e:
                    st.error(f'Erro ao adicionar a entrega: {e}')
            else:
                st.error('Digite os campos obrigatórios')